import json
import logging
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_EMPRESA_CONFIG = Path(__file__).parent / "empresa_urls.json"

# Final Excel column order: header label → row-dict key
# "estado", "fuente", "sector", "dias_en_bd" are computed fields injected at export time
COLUMNS = [
    ("Estado",     "estado"),
    ("Fecha",      "date_posted"),
    ("Días en BD", "dias_en_bd"),
    ("Fuente",     "fuente"),
    ("Empresa",    "company"),
    ("Puesto",     "title"),
    ("Sector",     "sector"),
    ("Ubicación",  "location"),
    ("Seniority",  "seniority"),
    ("Salario",    "salary"),
    ("URL",        "url"),
]

# Fields actually stored in the DB (computed fields are not here)
_DB_FIELDS = [
    "company", "title", "location", "salary",
    "date_posted", "seniority", "tipo_fuente", "url", "source",
    "fecha_ingreso_db",
]

# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
ALT_ROW_FILL = PatternFill(start_color="EBF0F8", end_color="EBF0F8", fill_type="solid")
LINK_FONT    = Font(color="1155CC", underline="single", name="Calibri", size=10)
BODY_FONT    = Font(name="Calibri", size=10)
THIN_BORDER  = Border(
    left=Side(style="thin",   color="D0D7E3"),
    right=Side(style="thin",  color="D0D7E3"),
    top=Side(style="thin",    color="D0D7E3"),
    bottom=Side(style="thin", color="D0D7E3"),
)

COL_WIDTHS = {
    "Estado":     13,
    "Fecha":      14,
    "Días en BD": 10,
    "Fuente":     18,
    "Empresa":    28,
    "Puesto":     42,
    "Sector":     26,
    "Ubicación":  22,
    "Seniority":  24,
    "Salario":    18,
    "URL":        60,
}

# ── Sector detection ──────────────────────────────────────────────────────────

_SECTOR_RULES = [
    ("Finanzas y Banca",      ["finan", "financiero", "cfo", "controller",
                               "tesorerí", "tesoreria", "banca", "credito", "crédito"]),
    ("Estrategia y Negocios", ["estrateg", "strateg", "business development",
                               "biz dev", "desarrollo de negocio", "planeamiento", "corporativ"]),
    ("Comercial",             ["comercial", "ventas", "marketing", "trade"]),
    ("Operaciones",           ["operacion", "supply chain", "logístic",
                               "logistic", "cadena de suministro"]),
    ("Tecnología",            ["tecnolog", "sistemas", "digital",
                               "software", " it ", " ti ", "datos", "data"]),
]
_SECTOR_DEFAULT = "Por determinar"


def _load_company_sectors() -> dict:
    """Return {company_name_lower: sector} from empresa_urls.json."""
    if not _EMPRESA_CONFIG.exists():
        return {}
    try:
        data = json.loads(_EMPRESA_CONFIG.read_text(encoding="utf-8"))
        return {
            e["nombre"].lower(): e.get("sector", _SECTOR_DEFAULT)
            for e in data.get("empresas", [])
        }
    except Exception as exc:
        logger.warning(f"Could not load company sectors: {exc}")
        return {}


def _detect_sector(title: str, company: str, tipo_fuente: str,
                   company_sectors: dict) -> str:
    if tipo_fuente == "Empresa Directa":
        return company_sectors.get(company.lower(), _SECTOR_DEFAULT)
    t = title.lower()
    for sector, tokens in _SECTOR_RULES:
        if any(tok in t for tok in tokens):
            return sector
    return _SECTOR_DEFAULT


def _get_fuente(source: str, tipo_fuente: str) -> str:
    if tipo_fuente == "Empresa Directa":
        return "Empresa Directa"
    return source or "Desconocido"


def _compute_dias_en_bd(fecha_ingreso_db: str) -> int | str:
    """Return integer days since the job was first saved; '' if unknown."""
    if not fecha_ingreso_db:
        return ""
    try:
        ingreso = datetime.fromisoformat(fecha_ingreso_db).date()
        return (date.today() - ingreso).days
    except Exception:
        return ""


def _compute_estado(dias_en_bd: int | str) -> str:
    if dias_en_bd == 0:
        return "🆕 Nuevo"
    return "📋 Visto"


# ── Public entry point ────────────────────────────────────────────────────────

def export_to_excel(conn, filename: str = "jobs_today.xlsx") -> int:
    today = date.today().isoformat()
    select_cols = ", ".join(_DB_FIELDS)

    # Fetch all jobs from the last 7 days (matches DB retention window)
    rows = conn.execute(
        f"SELECT {select_cols} FROM jobs "
        f"WHERE DATE(fecha_ingreso_db) >= DATE('now', '-6 days')",
    ).fetchall()

    if not rows:
        logger.warning("No jobs in DB for the last 7 days — Excel will be empty")

    company_sectors = _load_company_sectors()

    # Enrich each DB row with computed fields
    enriched: list[dict] = []
    for raw in rows:
        row = dict(zip(_DB_FIELDS, raw))
        row["fuente"]     = _get_fuente(row["source"], row.get("tipo_fuente") or "")
        row["sector"]     = _detect_sector(
            row["title"], row["company"],
            row.get("tipo_fuente") or "", company_sectors,
        )
        row["dias_en_bd"] = _compute_dias_en_bd(row.get("fecha_ingreso_db") or "")
        row["estado"]     = _compute_estado(row["dias_en_bd"])
        enriched.append(row)

    # Sort: Nuevo first, then by días en BD ascending (most recent first within each group)
    enriched.sort(key=lambda r: (
        0 if r["estado"] == "🆕 Nuevo" else 1,
        r["dias_en_bd"] if isinstance(r["dias_en_bd"], int) else 999,
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Empleos {today}"

    _write_main_sheet(ws, enriched, today)
    _write_summary_sheet(wb, enriched, today)

    wb.save(filename)
    logger.info(f"Saved {len(enriched)} jobs → {filename}")
    return len(enriched)


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_main_sheet(ws, rows: list[dict], today: str):
    header_labels = [c[0] for c in COLUMNS]
    ws.append(header_labels)

    # Style header row
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    url_col_idx    = next(i + 1 for i, (h, _) in enumerate(COLUMNS) if h == "URL")
    puesto_col_idx = next(i + 1 for i, (h, _) in enumerate(COLUMNS) if h == "Puesto")

    for row_num, row_data in enumerate(rows, start=2):
        row_values = [row_data.get(field, "") for _, field in COLUMNS]
        ws.append(row_values)

        for col_idx, cell in enumerate(ws[row_num], start=1):
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(col_idx == puesto_col_idx),
            )
            if col_idx == url_col_idx and cell.value:
                cell.hyperlink = str(cell.value)
                cell.font = LINK_FONT
            else:
                cell.font = BODY_FONT
                if row_num % 2 == 0:
                    cell.fill = ALT_ROW_FILL

        ws.row_dimensions[row_num].height = 18

    # Column widths
    for col_idx, (label, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[label]

    ws.auto_filter.ref = ws.dimensions


def _write_summary_sheet(wb, rows: list[dict], today: str):
    ws = wb.create_sheet("Resumen")

    title_font = Font(bold=True, size=14, color="1F3864", name="Calibri")
    bold_font  = Font(bold=True, name="Calibri")

    new_count = sum(1 for r in rows if r.get("estado") == "🆕 Nuevo")

    ws["A1"] = "Resumen de Búsqueda de Empleos — Perú"
    ws["A1"].font = title_font
    ws["A2"] = f"Fecha: {today}"
    ws["A3"] = f"Nuevos hoy: {new_count}"
    ws["A3"].font = bold_font
    ws["A4"] = f"Total activos (7 días): {len(rows)}"
    ws["A4"].font = bold_font

    ws.append([])

    def _section(title: str, counter: Counter):
        ws.append([title, "Cantidad"])
        for cell in ws[ws.max_row]:
            cell.font = bold_font
        for label, cnt in counter.most_common():
            ws.append([label, cnt])
        ws.append([])

    _section("Por fuente:",    Counter(r["fuente"]    for r in rows))
    _section("Por sector:",    Counter(r["sector"]    for r in rows))
    _section("Por seniority:", Counter(r["seniority"] for r in rows))

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
